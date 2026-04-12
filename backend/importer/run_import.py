"""
Excel Import Runner
===================
Reads Portfolio.xlsx, normalizes all sheets, and loads data into the database.

Usage (from backend/ directory):
    python -m importer.run_import
    python -m importer.run_import --file ../Portfolio.xlsx
    python -m importer.run_import --reset        # drop & recreate all tables, then import
    python -m importer.run_import --reset-data   # clear transactions/options/snapshots, then import

Import behaviour
----------------
Every section is idempotent. Running the importer multiple times is always safe:

  Holdings      No longer imported directly. Holdings are derived from transactions.
                The Holdings sheet is used only to seed bootstrap transactions on
                first import.

  Transactions  Inserted only when the natural key does not already exist:
                  (asset_id, type, trade_date, quantity, price_per_share)
                Duplicate rows in the source file are silently skipped.

  Options       Inserted only when the natural key does not already exist:
                  (underlying_asset_id, option_type, action, strike, expiration_date, open_date)
                Duplicate rows in the source file are silently skipped.

  Snapshots     Skipped entirely if a snapshot with the same date already exists.

Developer reset commands
------------------------
  --reset
      Drops ALL tables and recreates them from the ORM models, then runs a
      full import. Use when the schema itself has changed (e.g. a new column or
      constraint was added to the models). This is the safest way to apply
      schema changes to the local SQLite database without running a migration.

  --reset-data
      Deletes only transactional rows (transactions, option_trades, snapshots
      and their child snapshot_holdings). Assets and holdings are preserved.
      Use when you want a clean re-import of trading history without losing
      the holdings/asset catalogue, and without touching the schema.

  --rebuild-holdings
      Clear the holdings table and rebuild it deterministically from all
      transactions in the database. No Excel file is required. Runs integrity
      checks after rebuild and reports any discrepancies.

Natural key safety notes (MVP)
-------------------------------
The natural keys chosen are safe for the current dataset but have known limits:

  Transactions  Key: (asset_id, type, trade_date, quantity, price_per_share)
    SAFE:   Covers all normal BUY/SELL history from a single broker account.
    LIMIT:  Two genuinely separate trades on the same day, same stock, same
            quantity, and same price would be collapsed to one row. This is
            extremely unlikely in practice (partial fills are aggregated by
            most brokers before export) but theoretically possible.

  Option trades  Key: (underlying_asset_id, option_type, action, strike, expiration_date, open_date)
    SAFE:   Each option contract has a unique combination of underlying,
            type, action, strike, expiry, and open date in normal usage.
    LIMIT:  Rolling the same contract on the same day (buy-to-close + sell-to-
            open for identical strikes/expiry) would collapse if the new
            position is entered on the exact same day. This is rare and would
            only happen with same-day roll trades.

  Snapshot holdings  Key: (snapshot_id, asset_id)
    SAFE:   There is exactly one position per asset per year-end snapshot.
    No known limits for the current MVP.

If the dataset ever grows to include intraday or multi-leg option strategies,
revisit these keys and consider adding a broker-assigned trade ID column.
"""

import argparse
import logging
import sys
from pathlib import Path

import openpyxl
from sqlalchemy.orm import Session

# Make sure app modules are importable when run from backend/
sys.path.insert(0, str(Path(__file__).parent.parent))

from app.database import SessionLocal, create_tables, engine, Base
from app.config import settings
from importer.parsers import holdings_parser, transactions_parser, options_parser, snapshot_parser
from importer.loaders.db_loader import LoadResult, load_bootstrap_transactions, load_transactions, load_options, load_snapshot
from app.services.transaction_service import recalculate_all_holdings, rebuild_holdings

logging.basicConfig(level=logging.INFO, format="%(levelname)s  %(message)s")
logger = logging.getLogger(__name__)


def fetch_fx_rates(currencies: set[str]) -> dict[str, float]:
    """Fetch current FX rates for non-USD currencies using yfinance."""
    rates: dict[str, float] = {}
    non_usd = {c for c in currencies if c != "USD"}
    if not non_usd:
        return rates

    try:
        import yfinance as yf
        from app.services.market_data_service import FX_TICKERS
        for currency in non_usd:
            ticker_sym = FX_TICKERS.get(currency)
            if not ticker_sym:
                logger.warning("No FX ticker for %s — using 1.0", currency)
                rates[currency] = 1.0
                continue
            rate = yf.Ticker(ticker_sym).fast_info.last_price
            rates[currency] = float(rate) if rate else 1.0
            logger.info("FX rate %s/USD = %.4f", currency, rates[currency])
    except Exception as exc:
        logger.warning("Could not fetch FX rates: %s. Non-USD avg costs will be 1:1.", exc)
        for c in non_usd:
            rates[c] = 1.0
    return rates


def reset_transactional_data(db: Session) -> None:
    """
    Delete all transactions, option trades, and snapshots without touching
    assets or holdings. Use this to re-import cleanly without a full DB reset.
    """
    from app.models.transaction import Transaction
    from app.models.option_trade import OptionTrade
    from app.models.snapshot import PortfolioSnapshot

    deleted_tx = db.query(Transaction).delete()
    deleted_opt = db.query(OptionTrade).delete()
    deleted_snap = db.query(PortfolioSnapshot).delete()  # cascades to snapshot_holdings
    db.commit()
    logger.warning(
        "--reset-data: deleted %d transactions, %d option trades, %d snapshots.",
        deleted_tx, deleted_opt, deleted_snap,
    )


def _print_summary(
    bootstrap: LoadResult,
    transactions: LoadResult,
    options: LoadResult,
    snapshots: LoadResult,
) -> None:
    """Print a structured import summary to the log."""
    bar = "=" * 50
    logger.info(bar)
    logger.info("IMPORT SUMMARY")
    logger.info(bar)
    logger.info(
        "  Bootstrap tx : %3d inserted   %3d skipped   (%d total)",
        bootstrap.inserted, bootstrap.skipped, bootstrap.total,
    )
    logger.info(
        "  Transactions : %3d inserted   %3d skipped   (%d total)",
        transactions.inserted, transactions.skipped, transactions.total,
    )
    logger.info(
        "  Options      : %3d inserted   %3d skipped   (%d total)",
        options.inserted, options.skipped, options.total,
    )
    logger.info(
        "  Snapshots    : %3d inserted   %3d skipped   (%d total)",
        snapshots.inserted, snapshots.skipped, snapshots.total,
    )
    logger.info(bar)
    if transactions.skipped == transactions.total and transactions.total > 0:
        logger.info("  All transactions already existed — no duplicates were created.")
    if options.skipped == options.total and options.total > 0:
        logger.info("  All option trades already existed — no duplicates were created.")


def run(excel_path: str, reset: bool = False, reset_data: bool = False) -> None:
    path = Path(excel_path)
    if not path.exists():
        logger.error("Excel file not found: %s", path)
        sys.exit(1)

    logger.info("Opening workbook: %s", path)
    wb = openpyxl.load_workbook(str(path), data_only=True)

    if reset:
        logger.warning("--reset flag set: dropping all tables and recreating.")
        import app.models  # noqa: F401 — register all models
        Base.metadata.drop_all(bind=engine)

    create_tables()

    db = SessionLocal()
    try:
        if reset_data and not reset:
            reset_transactional_data(db)

        # ── 1. Bootstrap transactions from Holdings sheet ──────────────────────
        # Holdings sheet is used only to seed assets and bootstrap transactions.
        # It no longer directly upserts holdings — those are derived from transactions.
        bootstrap_result = LoadResult(inserted=0, skipped=0)
        raw_holdings = []
        fx_rates: dict[str, float] = {}
        if "Holdings" in wb.sheetnames:
            logger.info("Parsing Holdings sheet (migration source only)...")
            raw_holdings = holdings_parser.parse(wb["Holdings"])
            currencies = {rh.currency for rh in raw_holdings}
            fx_rates = fetch_fx_rates(currencies)
            logger.info("Seeding bootstrap transactions from holdings...")
            bootstrap_result = load_bootstrap_transactions(db, raw_holdings, fx_rates)
            db.commit()
        else:
            logger.warning("No 'Holdings' sheet found.")

        # ── 2. Transactions ────────────────────────────────────────────────────
        transactions_result = LoadResult(inserted=0, skipped=0)
        if "Transactions" in wb.sheetnames:
            logger.info("Parsing Transactions sheet...")
            raw_txs = transactions_parser.parse(wb["Transactions"])
            transactions_result = load_transactions(db, raw_txs, fx_rates)
            db.commit()
        else:
            logger.warning("No 'Transactions' sheet found.")

        # ── 2b. Recalculate all holdings from transactions ─────────────────────
        logger.info("Recalculating holdings from transactions...")
        recalculate_all_holdings(db)
        db.commit()

        # ── 3. Options ─────────────────────────────────────────────────────────
        options_result = LoadResult(inserted=0, skipped=0)
        if "Options" in wb.sheetnames:
            logger.info("Parsing Options sheet...")
            raw_opts = options_parser.parse(wb["Options"])
            options_result = load_options(db, raw_opts)
            db.commit()
        else:
            logger.warning("No 'Options' sheet found.")

        # ── 4. Snapshot_YYYY sheets (auto-detected) ───────────────────────────
        snapshots_result = LoadResult(inserted=0, skipped=0)
        snapshot_sheets = [s for s in wb.sheetnames if snapshot_parser.is_snapshot_sheet(s)]
        if snapshot_sheets:
            snapshot_sheets.sort(key=snapshot_parser.extract_year)
            for sheet_name in snapshot_sheets:
                logger.info("Parsing snapshot sheet: %s", sheet_name)
                raw_snap = snapshot_parser.parse(wb[sheet_name], sheet_name)
                result = load_snapshot(db, raw_snap, fx_rates)
                snapshots_result = LoadResult(
                    inserted=snapshots_result.inserted + result.inserted,
                    skipped=snapshots_result.skipped + result.skipped,
                )
                db.commit()
        else:
            logger.info("No Snapshot_YYYY sheets found.")

        _print_summary(bootstrap_result, transactions_result, options_result, snapshots_result)

    except Exception as exc:
        db.rollback()
        logger.error("Import failed: %s", exc)
        raise
    finally:
        db.close()


def main():
    parser = argparse.ArgumentParser(description="Import Portfolio.xlsx into the database.")
    parser.add_argument("--file", default=settings.excel_file_path, help="Path to Portfolio.xlsx")
    parser.add_argument("--reset", action="store_true", help="Drop and recreate all tables before import")
    parser.add_argument(
        "--reset-data", action="store_true", dest="reset_data",
        help="Clear transactions, option trades, and snapshots before import (keeps assets/holdings)",
    )
    parser.add_argument(
        "--rebuild-holdings", action="store_true", dest="rebuild_holdings",
        help="Clear and rebuild the holdings table from transaction history (no Excel file required)",
    )
    args = parser.parse_args()

    if args.rebuild_holdings:
        create_tables()
        db = SessionLocal()
        try:
            logger.info("Rebuilding holdings from transaction history...")
            result = rebuild_holdings(db)
            db.commit()
            bar = "=" * 50
            logger.info(bar)
            logger.info("REBUILD SUMMARY")
            logger.info(bar)
            logger.info("  Cleared  : %d holdings", result["cleared"])
            logger.info("  Rebuilt  : %d holdings", result["rebuilt"])
            if result["valid"]:
                logger.info("  Integrity: PASSED")
            else:
                logger.warning("  Integrity: FAILED")
                for err in result["errors"]:
                    logger.warning("    - %s", err)
            logger.info(bar)
        except Exception as exc:
            db.rollback()
            logger.error("Rebuild failed: %s", exc)
            raise
        finally:
            db.close()
        return

    run(args.file, reset=args.reset, reset_data=args.reset_data)


if __name__ == "__main__":
    main()
